"""Acceptance tests against the real workbook fixtures/JW_MRP_2026.xlsm."""
from __future__ import annotations

import random
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from mrp_assistant import db as dbmod
from mrp_assistant.calc import compute_line
from mrp_assistant.config import DEFAULTS
from mrp_assistant.diff import compute_diff
from mrp_assistant.ingest import ingest_file
from mrp_assistant.parsing import (
    parse_tier_qty,
    parse_tier_sheet,
    parse_weekly_sheet,
    parse_workbook,
)

FIXTURE = Path(__file__).resolve().parent.parent / "fixtures" / "JW_MRP_2026.xlsm"

pytestmark = pytest.mark.skipif(
    not FIXTURE.exists(), reason="real workbook fixture not present"
)


@pytest.fixture(scope="module")
def workbook():
    wb = openpyxl.load_workbook(FIXTURE, data_only=True, read_only=True)
    yield wb
    wb.close()


@pytest.fixture(scope="module")
def sheet_0529(workbook):
    return parse_weekly_sheet(workbook["05.29.26"], "05.29.26")


def test_1_tab_0529_snapshot_and_row_count(sheet_0529, workbook):
    """Tab 05.29.26: snapshot date 2026-05-29; all data rows survive
    subtotal-row filtering.

    NOTE: the spec asked for >250 rows, but the sheet's own "Grand Count"
    subtotal says 202 and manual inspection agrees — we trust the sheet.
    """
    assert sheet_0529.snapshot_date == date(2026, 5, 29)
    grand_count = None
    for row in workbook["05.29.26"].iter_rows(values_only=True):
        if row[1] == "Grand Count":
            grand_count = row[2]
    assert grand_count == 202
    assert len(sheet_0529.lines) == grand_count
    assert len(sheet_0529.lines) > 150
    # No subtotal rows leaked through.
    assert all("count" not in ln.item_number.lower() for ln in sheet_0529.lines)
    assert all(ln.msg_typ for ln in sheet_0529.lines)
    # Whitespace trimmed (Demand Branch is padded in the raw sheet).
    assert all(
        ln.demand_branch == ln.demand_branch.strip()
        for ln in sheet_0529.lines if ln.demand_branch
    )


def test_2_item_10105465_matches_workbook_p_q_columns(sheet_0529, workbook):
    """Item 10105465 (O, lead 190, request 2027-06-01, QA 60): projected
    arrival as-of-snapshot = 2026-12-05, slack = 118 — matching the
    workbook's own P/Q columns exactly."""
    line = next(ln for ln in sheet_0529.lines if ln.item_number == "10105465")
    assert line.msg_typ == "O"
    assert line.lead_time == 190
    assert line.request_date == date(2027, 6, 1)
    assert line.qa_days == 60

    computed = compute_line(
        {
            "msg_typ": line.msg_typ,
            "request_date": line.request_date,
            "qa_days": line.qa_days,
            "lead_time": line.lead_time,
        },
        snapshot_date=sheet_0529.snapshot_date,
        today=date.today(),
        cfg=DEFAULTS,
    )
    assert computed["snapshot"]["projected_dock_date"] == date(2026, 12, 5)
    assert computed["snapshot"]["slack_days"] == 118
    assert computed["effective_deadline"] == date(2027, 4, 2)
    assert computed["last_safe_order_date"] == date(2026, 9, 24)

    # Cross-check against the workbook's own P/Q cells for that row.
    ws = workbook["05.29.26"]
    wb_row = next(
        r for r in ws.iter_rows(values_only=True)
        if r[0] == "O" and r[1] == 10105465
    )
    assert wb_row[15].date() == computed["snapshot"]["projected_dock_date"]
    assert wb_row[16] == computed["snapshot"]["slack_days"]


def test_3_tier_parser(workbook):
    assert parse_tier_qty("25k") == 25000
    assert parse_tier_qty("110k") == 110000
    assert parse_tier_qty("1,000k") == 1000000
    assert parse_tier_qty("TBD") is None
    assert parse_tier_qty("VARIOUS") is None

    tier_rows, _ = parse_tier_sheet(workbook["Standard Cost Tiers"])
    tbd = [t for t in tier_rows if t.tier_str == "TBD"]
    assert tbd, "expected TBD tier rows in the real sheet"
    assert all(t.incomplete for t in tbd)
    assert all("unparseable tier" in t.incomplete_reason for t in tbd)
    # VARIOUS item rows are stored (not dropped) and flagged incomplete.
    various = [t for t in tier_rows if t.item_no == "VARIOUS"]
    assert various and all(t.incomplete for t in various)
    # A known parseable ladder survives intact.
    farevabio = [t for t in tier_rows if t.item_no == "10235072"]
    assert sorted(t.tier_qty for t in farevabio) == [25000, 50000, 100000, 200000]


def test_4_duplicate_ingest_is_idempotent(tmp_path):
    db_path = tmp_path / "mrp.db"
    archive = tmp_path / "archive"
    first = ingest_file(FIXTURE, db_path=db_path, archive_dir=archive)
    assert first.status == "ingested"
    second = ingest_file(FIXTURE, db_path=db_path, archive_dir=archive)
    assert second.status == "duplicate, skipped"

    assert len(list(archive.glob("*.xlsm"))) == 1  # exactly one archive copy
    conn = dbmod.connect(db_path)
    uploads = conn.execute("SELECT * FROM uploads ORDER BY id").fetchall()
    assert len(uploads) == 2
    assert uploads[0]["status"] == "ingested"
    assert uploads[1]["status"] == "duplicate, skipped"
    # No second copy of the lines.
    n_snapshots = conn.execute("SELECT COUNT(*) AS n FROM snapshots").fetchone()["n"]
    assert n_snapshots == 1
    conn.close()


def test_5_diff_between_0522_and_0529(tmp_path):
    db_path = tmp_path / "mrp.db"
    ingest_file(FIXTURE, tab="05.22.26", db_path=db_path, archive_dir=tmp_path / "a")
    result = ingest_file(
        FIXTURE, tab="05.29.26", db_path=db_path, archive_dir=tmp_path / "a", force=True
    )
    assert result.snapshots, "05.29.26 snapshot should have been created"
    new_id = result.snapshots[0]["snapshot_id"]

    conn = dbmod.connect(db_path)
    old_id = conn.execute(
        "SELECT id FROM snapshots WHERE source_tab = '05.22.26'"
    ).fetchone()["id"]
    rows = compute_diff(conn, new_id, old_id)
    kinds = {r["kind"] for r in rows}
    assert {"NEW", "RESOLVED", "CHANGED"} <= kinds

    # Every diff row is traceable to source rows with the same key.
    for r in rows:
        line_ids = r["new_line_ids"] + r["old_line_ids"]
        assert line_ids
        for line_id in line_ids:
            src = conn.execute(
                "SELECT * FROM mrp_lines WHERE id = ?", (line_id,)
            ).fetchone()
            assert src is not None
            assert src["item_number"] == r["item_number"]
            assert src["supplier"] == r["supplier"]
            assert src["request_date"] == r["request_date"]

    # The on-ingest stored diff matches the recomputed one.
    stored = conn.execute(
        "SELECT COUNT(*) AS n FROM diffs WHERE snapshot_id = ?", (new_id,)
    ).fetchone()["n"]
    assert stored == len(rows)
    conn.close()


def test_6_header_order_shuffled_copy_parses(workbook, tmp_path):
    """Column detection is by header name: a column-shuffled copy of the
    weekly tab must parse identically."""
    src = workbook["05.29.26"]
    rows = list(src.iter_rows(values_only=True))
    n_cols = max(len(r) for r in rows)
    order = list(range(n_cols))
    random.Random(42).shuffle(order)

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "05.29.26"
    for row in rows:
        padded = list(row) + [None] * (n_cols - len(row))
        ws.append([padded[i] for i in order])
    shuffled_path = tmp_path / "shuffled.xlsx"
    out.save(shuffled_path)

    original = parse_weekly_sheet(src, "05.29.26")
    parsed = parse_workbook(shuffled_path)
    assert len(parsed.weekly_sheets) == 1
    shuffled = parsed.weekly_sheets[0]
    assert shuffled.snapshot_date == date(2026, 5, 29)
    assert len(shuffled.lines) == len(original.lines)
    a = next(ln for ln in shuffled.lines if ln.item_number == "10105465")
    b = next(ln for ln in original.lines if ln.item_number == "10105465")
    for f in ("msg_typ", "lead_time", "request_date", "required_qty",
              "qa_days", "supplier", "supplier_name", "demand_branch"):
        assert getattr(a, f) == getattr(b, f)

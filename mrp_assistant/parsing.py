"""Workbook parsing: weekly MRP tabs (header-name driven), Standard Cost Tiers,
and SUPPLIERS reference sheets.

All values are read with data_only=True (computed values, never formulas) and
every derived field is recomputed in Python — workbook formulas may be stale
or #VALUE!.
"""
from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl

# Tab names like "05.29.26", "2.20.26", "01.02.26" (also tolerate - or _).
# Full-match only, so "Data 02.06.25" and other prefixed tabs are excluded.
TAB_DATE_RE = re.compile(r"^\s*(\d{1,2})[._-](\d{1,2})[._-](\d{2,4})\s*$")

# Canonical field name by normalized header text. Detection is ALWAYS by
# header name, never by position — column count drifts week to week.
WEEKLY_HEADER_MAP = {
    "MSG TYP": "msg_typ",
    "ITEM NUMBER": "item_number",
    "DESCRIPTION": "description",
    "LEAD TIME": "lead_time",
    "QA REQUEST DATE": "qa_request_date",
    "ACTUAL REQUEST DATE": "actual_request_date",
    "REQUEST DATE": "request_date",
    "REQUIRED QUANTITY": "required_qty",
    "ORDER NUMBER": "order_number",
    "SUPPLIER": "supplier",
    "SUPPLIER NAME": "supplier_name",
    "MESSAGE": "message",
    "OR TY": "or_ty",
    "QA DAYS": "qa_days",
    "DEMAND BRANCH": "demand_branch",
}

REQUIRED_WEEKLY_FIELDS = {
    "msg_typ",
    "item_number",
    "lead_time",
    "request_date",
    "required_qty",
    "qa_days",
}


@dataclass
class WeeklyLine:
    row_id: int  # 1-based source row number in the sheet, for traceability
    msg_typ: str
    item_number: str
    description: str | None = None
    lead_time: int | None = None
    qa_request_date: date | None = None
    actual_request_date: date | None = None
    request_date: date | None = None
    required_qty: float | None = None
    order_number: str | None = None
    supplier: str | None = None
    supplier_name: str | None = None
    message: str | None = None
    or_ty: str | None = None
    qa_days: int | None = None
    demand_branch: str | None = None


@dataclass
class ParsedSheet:
    tab: str
    snapshot_date: date
    tab_date: date | None
    header_date: date | None
    lines: list[WeeklyLine] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    skipped_rows: int = 0

    @property
    def content_hash(self) -> str:
        """Hash of the parsed line content, used to skip unchanged re-ingests."""
        payload = json.dumps(
            [
                [
                    ln.msg_typ, ln.item_number, ln.description, ln.lead_time,
                    _iso(ln.qa_request_date), _iso(ln.actual_request_date),
                    _iso(ln.request_date), ln.required_qty, ln.order_number,
                    ln.supplier, ln.supplier_name, ln.message, ln.or_ty,
                    ln.qa_days, ln.demand_branch,
                ]
                for ln in self.lines
            ],
            default=str,
        )
        return hashlib.sha256(payload.encode()).hexdigest()


@dataclass
class TierRow:
    supplier_name: str | None
    supplier_no: str | None
    item_no: str | None
    item_name: str | None
    tier_str: str | None
    tier_qty: int | None
    pricing: float | None
    pricing_year: int | None
    notes: str | None
    incomplete: bool
    incomplete_reason: str | None


@dataclass
class SupplierRow:
    coo: str | None
    name: str | None
    number: str | None
    branches: list[str]


@dataclass
class ParsedWorkbook:
    path: str
    weekly_sheets: list[ParsedSheet]
    tier_rows: list[TierRow]
    suppliers: list[SupplierRow]
    warnings: list[str]


def _iso(d: date | None) -> str | None:
    return d.isoformat() if d else None


def _norm_header(value) -> str:
    return str(value).strip().upper() if value is not None else ""


def _to_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _to_int(value) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        try:
            return int(float(text))
        except ValueError:
            return None
    return None


def _to_float(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def parse_tab_name_date(tab_name: str) -> date | None:
    """Parse MM.DD.YY style tab names; returns None for non-date tabs."""
    m = TAB_DATE_RE.match(tab_name)
    if not m:
        return None
    month, day, year = (int(g) for g in m.groups())
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_tier_qty(tier_str) -> int | None:
    """Parse tier strings: "25k" -> 25000, "1,000k" -> 1000000, "TBD" -> None."""
    if tier_str is None:
        return None
    if isinstance(tier_str, (int, float)) and not isinstance(tier_str, bool):
        return int(tier_str)
    text = str(tier_str).strip().lower().replace(",", "")
    if not text:
        return None
    multiplier = 1
    if text.endswith("k"):
        multiplier = 1000
        text = text[:-1]
    try:
        return int(float(text) * multiplier)
    except ValueError:
        return None


def parse_weekly_sheet(ws, tab_name: str) -> ParsedSheet | None:
    """Parse one dated weekly tab. Returns None if the header row is not an
    MRP message table."""
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return None

    col_for: dict[str, int] = {}
    for idx, cell in enumerate(header):
        canon = WEEKLY_HEADER_MAP.get(_norm_header(cell))
        if canon and canon not in col_for:
            col_for[canon] = idx

    missing = REQUIRED_WEEKLY_FIELDS - col_for.keys()
    if missing:
        return None

    warnings: list[str] = []

    # Snapshot date: the cell to the right of the "Demand Branch" header.
    header_date = None
    db_idx = col_for.get("demand_branch")
    if db_idx is not None and db_idx + 1 < len(header):
        header_date = _to_date(header[db_idx + 1])
    if header_date is None:
        # Fall back to any date-looking value in the header row.
        for idx, cell in enumerate(header):
            if idx not in col_for.values() and isinstance(cell, (datetime, date)):
                header_date = _to_date(cell)
                break

    tab_date = parse_tab_name_date(tab_name)
    if header_date is None and tab_date is None:
        return None
    if header_date is None:
        warnings.append(
            f"[{tab_name}] no snapshot date in header row; using tab name date {tab_date}"
        )
    elif tab_date and header_date != tab_date:
        warnings.append(
            f"[{tab_name}] snapshot date mismatch: header cell says {header_date}, "
            f"tab name says {tab_date}; using header cell value"
        )
    snapshot_date = header_date or tab_date

    def cell(row, fieldname):
        idx = col_for.get(fieldname)
        return row[idx] if idx is not None and idx < len(row) else None

    lines: list[WeeklyLine] = []
    skipped = 0
    for row_no, row in enumerate(rows, start=2):
        msg_typ = _to_text(cell(row, "msg_typ"))
        raw_item = cell(row, "item_number")
        # Data hygiene: skip subtotal rows ("10094876 Count", "Grand Count",
        # SUBTOTAL leftovers) and anything with a blank Msg Typ.
        if not msg_typ:
            if raw_item is not None or any(v is not None for v in row):
                skipped += 1
            continue
        item_text = _to_text(raw_item) if not isinstance(raw_item, (int, float)) else str(_to_int(raw_item))
        if item_text is None or "count" in item_text.lower():
            skipped += 1
            continue

        line = WeeklyLine(
            row_id=row_no,
            msg_typ=msg_typ,
            item_number=item_text,
            description=_to_text(cell(row, "description")),
            lead_time=_to_int(cell(row, "lead_time")),
            qa_request_date=_to_date(cell(row, "qa_request_date")),
            actual_request_date=_to_date(cell(row, "actual_request_date")),
            request_date=_to_date(cell(row, "request_date")),
            required_qty=_to_float(cell(row, "required_qty")),
            order_number=_to_text(cell(row, "order_number")),
            supplier=_to_text(cell(row, "supplier")),
            supplier_name=_to_text(cell(row, "supplier_name")),
            message=_to_text(cell(row, "message")),
            or_ty=_to_text(cell(row, "or_ty")),
            qa_days=_to_int(cell(row, "qa_days")),
            demand_branch=_to_text(cell(row, "demand_branch")),
        )
        for fieldname in ("lead_time", "request_date", "required_qty", "qa_days"):
            if getattr(line, fieldname) is None:
                raw = cell(row, fieldname)
                warnings.append(
                    f"[{tab_name}] row {row_no} item {item_text}: "
                    f"unreadable {fieldname} ({raw!r})"
                )
        lines.append(line)

    return ParsedSheet(
        tab=tab_name,
        snapshot_date=snapshot_date,
        tab_date=tab_date,
        header_date=header_date,
        lines=lines,
        warnings=warnings,
        skipped_rows=skipped,
    )


def parse_tier_sheet(ws) -> tuple[list[TierRow], list[str]]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return [], ["Standard Cost Tiers sheet is empty"]
    col_map = {
        "SUPPLIER NAME": "supplier_name",
        "SUPPLIER NO.": "supplier_no",
        "ITEM NO.": "item_no",
        "ITEM NAME": "item_name",
        "ORDERING TIERS": "tier_str",
        "PRICING": "pricing",
        "PRICING YEAR": "pricing_year",
        "NOTES": "notes",
    }
    col_for = {}
    for idx, cell in enumerate(header):
        canon = col_map.get(_norm_header(cell))
        if canon and canon not in col_for:
            col_for[canon] = idx
    warnings: list[str] = []
    out: list[TierRow] = []

    def cell(row, fieldname):
        idx = col_for.get(fieldname)
        return row[idx] if idx is not None and idx < len(row) else None

    for row_no, row in enumerate(rows, start=2):
        supplier_name = _to_text(cell(row, "supplier_name"))
        item_no_raw = cell(row, "item_no")
        if supplier_name is None and item_no_raw is None:
            continue
        tier_str_raw = cell(row, "tier_str")
        tier_str = _to_text(tier_str_raw)
        tier_qty = parse_tier_qty(tier_str_raw)
        pricing = _to_float(cell(row, "pricing"))
        item_no = (
            str(_to_int(item_no_raw))
            if isinstance(item_no_raw, (int, float))
            else _to_text(item_no_raw)
        )
        # Junk values ("TBD", "VARIOUS") are stored but flagged incomplete,
        # never silently dropped.
        reasons = []
        if tier_qty is None:
            reasons.append(f"unparseable tier {tier_str!r}")
        if item_no is not None and item_no.upper() == "VARIOUS":
            reasons.append("item listed as VARIOUS")
        if pricing is None:
            reasons.append(f"unparseable price {_to_text(cell(row, 'pricing'))!r}")
        out.append(
            TierRow(
                supplier_name=supplier_name,
                supplier_no=str(_to_int(cell(row, "supplier_no")) or _to_text(cell(row, "supplier_no")) or ""),
                item_no=item_no,
                item_name=_to_text(cell(row, "item_name")),
                tier_str=tier_str,
                tier_qty=tier_qty,
                pricing=pricing,
                pricing_year=_to_int(cell(row, "pricing_year")),
                notes=_to_text(cell(row, "notes")),
                incomplete=bool(reasons),
                incomplete_reason="; ".join(reasons) if reasons else None,
            )
        )
        if reasons:
            warnings.append(
                f"[Standard Cost Tiers] row {row_no} "
                f"({supplier_name} / {item_no}): tier data incomplete: "
                + "; ".join(reasons)
            )
    return out, warnings


def parse_suppliers_sheet(ws) -> list[SupplierRow]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []
    name_idx = number_idx = coo_idx = None
    branch_idxs = []
    for idx, cell in enumerate(header):
        norm = _norm_header(cell)
        if norm == "COO":
            coo_idx = idx
        elif norm == "SUPPLIER NAME":
            name_idx = idx
        elif norm == "SUPPLIER NUMBER":
            number_idx = idx
        elif norm.startswith("BRANCH"):
            branch_idxs.append(idx)
    out = []
    for row in rows:
        name = _to_text(row[name_idx]) if name_idx is not None and name_idx < len(row) else None
        if not name:
            continue
        number_raw = row[number_idx] if number_idx is not None and number_idx < len(row) else None
        number = (
            str(_to_int(number_raw))
            if isinstance(number_raw, (int, float))
            else _to_text(number_raw)
        )
        coo = _to_text(row[coo_idx]) if coo_idx is not None and coo_idx < len(row) else None
        branches = []
        for idx in branch_idxs:
            if idx < len(row):
                b = _to_text(row[idx])
                if b:
                    branches.append(b)
        out.append(SupplierRow(coo=coo, name=name, number=number, branches=branches))
    return out


def parse_workbook(path: str | Path) -> ParsedWorkbook:
    """Parse every dated weekly tab plus the reference sheets."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        weekly: list[ParsedSheet] = []
        warnings: list[str] = []
        tier_rows: list[TierRow] = []
        suppliers: list[SupplierRow] = []
        for name in wb.sheetnames:
            if parse_tab_name_date(name):
                sheet = parse_weekly_sheet(wb[name], name)
                if sheet is None:
                    warnings.append(f"[{name}] looks like a dated tab but has no MRP header row; skipped")
                else:
                    weekly.append(sheet)
                    warnings.extend(sheet.warnings)
            elif _norm_header(name) == "STANDARD COST TIERS":
                tier_rows, tier_warnings = parse_tier_sheet(wb[name])
                warnings.extend(tier_warnings)
            elif _norm_header(name) == "SUPPLIERS":
                suppliers = parse_suppliers_sheet(wb[name])
        weekly.sort(key=lambda s: s.snapshot_date)
        return ParsedWorkbook(
            path=str(path),
            weekly_sheets=weekly,
            tier_rows=tier_rows,
            suppliers=suppliers,
            warnings=warnings,
        )
    finally:
        wb.close()

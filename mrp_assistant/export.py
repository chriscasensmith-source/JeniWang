"""Excel exports: current worksheet view and the week-over-week diff,
stamped with snapshot date and generation timestamp."""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FILLS = {
    "RED": PatternFill("solid", fgColor="FFC7CE"),
    "YELLOW": PatternFill("solid", fgColor="FFEB9C"),
    "GREEN": PatternFill("solid", fgColor="E2EFDA"),
}

WORKSHEET_COLUMNS = [
    ("Msg Typ", lambda l, m: l["msg_typ"]),
    ("Item Number", lambda l, m: l["item_number"]),
    ("Description", lambda l, m: l["description"]),
    ("Supplier", lambda l, m: l["supplier"]),
    ("Supplier Name", lambda l, m: l["supplier_name"]),
    ("Lead Time", lambda l, m: l["lead_time"]),
    ("QA Days", lambda l, m: l["qa_days"]),
    ("Request Date", lambda l, m: l["request_date"]),
    ("Effective Deadline", lambda l, m: l["effective_deadline"]),
    ("Last Safe Order Date", lambda l, m: l["last_safe_order_date"]),
    ("Order Date Used", lambda l, m: l[m]["order_date"]),
    ("Projected Dock Date", lambda l, m: l[m]["projected_dock_date"]),
    ("Slack Days", lambda l, m: l[m]["slack_days"]),
    ("Status", lambda l, m: l[m]["status"]),
    ("Required Qty", lambda l, m: l["required_qty"]),
    ("Order Number", lambda l, m: l["order_number"]),
    ("Demand Branch", lambda l, m: l["demand_branch"]),
    ("Source Row", lambda l, m: l["row_id"]),
]


def _stamp(ws, title: str, snapshot_date: str) -> None:
    ws.append([title])
    ws.append([
        f"Snapshot: {snapshot_date}",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
    ])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=12)


def worksheet_xlsx(board: dict, mode: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    snapshot_date = board["snapshot"]["snapshot_date"]
    _stamp(ws, f"MRP worksheet view (order date as of {mode})", snapshot_date)
    ws.append([name for name, _ in WORKSHEET_COLUMNS])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for line in board["lines"]:
        ws.append([getter(line, mode) for _, getter in WORKSHEET_COLUMNS])
        status = line[mode]["status"]
        if status in FILLS:
            for cell in ws[ws.max_row]:
                cell.fill = FILLS[status]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def diff_xlsx(board: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Diff"
    snapshot_date = board["snapshot"]["snapshot_date"]
    diff = board.get("diff")
    prev = diff["prev_snapshot_date"] if diff else "n/a"
    _stamp(ws, f"Week-over-week diff: {prev} -> {snapshot_date}", snapshot_date)
    header = ["Kind", "Item Number", "Supplier", "Supplier Name", "Request Date",
              "Old Qty", "New Qty", "Old Source Line IDs", "New Source Line IDs"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    if diff:
        for r in diff["rows"]:
            ws.append([
                r["kind"], r["item_number"], r["supplier"], r["supplier_name"],
                r["request_date"], r["old_qty"], r["new_qty"],
                ", ".join(map(str, r["old_line_ids"])),
                ", ".join(map(str, r["new_line_ids"])),
            ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
